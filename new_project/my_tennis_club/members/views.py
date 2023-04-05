from django.shortcuts import render
#from members.models import Productfield
#from tablib import Dataset
import pandas as pd
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import PatternFill
from django.core.paginator import Paginator
from django.shortcuts import render



# Create your views here.

'''def show(request):
    Products = Productfield.objects.all()
    i=0
    for product in Products:
        Products[i].Price_Per_100=float(product.Price_Per_100)+float(0.50)
        i=i+1
    # return HttpResponse(Product[0].Price_Per_100)
    return render(request,"show.html",{'Products':Products})'''

from django.http import HttpResponse  
from members.functions.functions import handle_uploaded_file  
from members.forms import StudentForm  
def index(request):  
    if request.method == 'POST':  
        student = StudentForm(request.POST, request.FILES)  
        if student.is_valid():  
            handle_uploaded_file(request.FILES['file'])  
            return HttpResponse("File uploaded successfuly")  
    else:  
        student = StudentForm()  
        return render(request,"index.html",{'form':student})  


import datetime
#import math
#import pymysql
#from django.core.paginator import Paginator
from urllib.parse import quote_plus


#creating function for uplode data into database

def simple_upload(request):
    
    #save ip address function call
    show_ip_address(request)
    
    # get uploaded file

    if request.method=="POST":
        now=datetime.datetime.now()
        current_date=now.strftime('%Y-%m-%d')
        engine=create_engine("mysql+pymysql://root:%s@localhost/indicative" % quote_plus("Pranay@123"))
        #print("p")
        new_person=request.FILES["myfile"]
        dataframe=pd.read_excel(new_person)
        l=len(dataframe)
        l=l+2
        #print(l)
        dataframe.rename(columns={'Maturity Date ':'Maturity Date'},inplace=True)
        dataframe.rename(columns={'Put/Call Option ':'Put/Call Option'},inplace=True)

        #print(dataframe)
        dataframe.to_excel("C:\\project1\\new_project\\my_tennis_club\\members\\static\\daily_data\\{}_{}".format(current_date,new_person))

        dataframe.to_excel("C:\\project1\\new_project\\my_tennis_club\\members\\static\\upload\\{}".format('Indicative.xlsx'))
        #data=['Coupon Rate','ISIN','Name of the Security','CATEGORY','Rating & Agency','Maturity Date','IP Dates','Put/Call Option ']
        dataframe.to_sql('data1',con=engine,if_exists="replace",index=False)
        #table=dataframe.to_html()
        #print(table)

        # get price from html code

        num=request.POST.get('num')
        #print(num)
        
        

        #print(table('ISIN'))
        #Products = table.objects.all()
        #i=0
        #for product in Products:
           #[i].Price_Per_100=float(product.Price_Per_100)+float(0.50)
           #i=i+1
        calculation(l,num,current_date)   
        #new_person=request.FILES["C:\project\website\hello_world.xlsx"]
        # Read html file

        dataframe=pd.read_excel("C:\\project1\\new_project\\my_tennis_club\\hello_world.xlsx")
        dataframe.to_excel("C:\\project1\\new_project\\my_tennis_club\\members\\static\\calculated_data\\{}_{}".format(current_date,new_person))
        
        table=dataframe.to_html()
        #print(table)
        
        return render(request,'members/index.html',{'table':table})
    else:
       return render(request,'members/index.html')

#for pegination

def my_view(request):
    # Load the Excel file
    workbook = openpyxl.load_workbook('C:\\project1\\new_project\\my_tennis_club\\hello_world.xlsx')
    worksheet = workbook.active
    print(worksheet)

    rows = worksheet.rows
    # Convert the Excel data to a list of dictionaries
    data = []
    columns = [cell.value for cell in next(rows)]
    for row in rows:
        data.append({columns[i]: cell.value for i, cell in enumerate(row)})
        # Paginate the data
    paginator = Paginator(data, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
        # Pass the paginated data to the template
    return render(request, 'pegination.html', {'page_obj': page_obj})



    
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter





def calculation(l,num,current_date):
    
    #get path from saved file
    path = "members\\static\\upload\\Indicative.xlsx"
    now=datetime.datetime.now()
    current_date=now.strftime('%d%b-%Y').replace('-', '-')
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    #print(sheet_obj['C2'].value)
    sheet_obj.insert_rows(idx=2)
    l=l+1
    sheet_obj.merge_cells('B2:P2') 
    border(sheet_obj,2)
    sheet_obj["B2"]=f"Indicative Quotes-{current_date}"
    sheet_obj["B2"].font = Font(size=18,bold=True)
    '''pad = Padding(top=5, bottom=5)
    sheet_obj['B2'].Padding = pad'''


    #sheet_obj.insert_rows(idx=3)
    #l=l+1
    #sheet_obj[cell]='P' 
    #cell = sheet_obj.cell(row=2, column=2)  

    #cell.value = 'Devansh Sharma'  
    sheet_obj["B2"].alignment = Alignment(horizontal='center', vertical='center')  

    #Category : GOI & SDL BONDS(NSDL DP Only)
    #sheet_obj.insert_rows(idx=3)
    #l=l+1
    #sheet_obj.merge_cells('B3:P3') 
    #sheet_obj["B3"]="Category : GOI & SDL BONDS(NSDL DP Only)"
    #sheet_obj['B3'].font = Font(bold=True)
    #sheet_obj["B3"].alignment = Alignment(horizontal='center', vertical='center')  
    #sheet_obj["A3"].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
    #print(sheet_obj["Q5"].value)
    
    '''h=["GOI",'PSU_TAX','PSU_PERPUTUAL','PSU_BOUND','State_Guaranteed','Private_Sector_AAA','Private_Sector_Bonds','Private_Sector_Perpetual_Bonds']
    for j in h:
        for i in range(1,l+2):
            if sheet_obj[f"B{i}"].value==j:
                sheet_obj.insert_rows(idx=i+1)
                sheet_obj.insert_rows(idx=i-1)'''
                




    # calculated price 
    for i in range(1,l):
        if sheet_obj[f"k{i}"].value==0:
            sheet_obj[f"J{i}"] =sheet_obj[f"J{i}"].value+float(num)
            border1(sheet_obj,i)
    # calculated Ytm,YTc and Yield
    for i in range(1,l):
        if sheet_obj[f"k{i}"].value==0:
        #print(sheet_obj[f'P{i}'].value)
        #print(sheet_obj[f'I{i}'].value)

            if sheet_obj[f'G{i}'].value !=0 and sheet_obj[f"I{i}"].value !=0:
                #print("p")
                if sheet_obj[f"P{i}"].value==2:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,G{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"L{i}"]=f"=EFFECT(K{i},2)"
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f"=EFFECT(K{i},2)"
                elif sheet_obj[f"P{i}"].value==4:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,G{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"L{i}"]=f"=EFFECT(K{i},4)" 
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f"=EFFECT(K{i},4)"    
                elif sheet_obj[f"P{i}"].value==12:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,G{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"L{i}"]=f"=EFFECT(K{i},12)" 
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f"=EFFECT(K{i},12)"       
                else:
                    sheet_obj[f"L{i}"]=f"=YIELD(C1,G{i},B{i},J{i},100,1,0)" 
                    sheet_obj[f"M{i}"]=f"=YIELD(C1,I{i},B{i},J{i},100,1,0)"
            
            elif sheet_obj[f'G{i}'].value !=0  :
                #print("p")

                #print((sheet_obj[f'P{i}'].value))
                if sheet_obj[f"P{i}"].value==2:
                    sheet_obj[f"K{i}"]=f'=YIELD($C$1,G{i},B{i},J{i},100,2,4)'
                    sheet_obj[f"L{i}"]=f'=EFFECT(K{i},2)'
                elif sheet_obj[f"P{i}"].value==4:
                    sheet_obj[f"K{i}"]=f'=YIELD($C$1,G{i},B{i},J{i},100,2,4)'
                    sheet_obj[f"L{i}"]=f'=EFFECT(K{i},4)'  
                elif sheet_obj[f"P{i}"].value==12:
                    sheet_obj[f"K{i}"]=f'=YIELD($C$1,G{i},B{i},J{i},100,2,4)'
                    sheet_obj[f"L{i}"]=f'=EFFECT(K{i},12)'  
                
                else:
                    sheet_obj[f"L{i}"]=f'=YIELD(C1,G{i},B{i},J{i},100,1,0)'  
            else:
                if sheet_obj[f"P{i}"].value==2:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f'=EFFECT(K{i},2)'
                elif sheet_obj[f"P{i}"].value==4:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f'=EFFECT(K{i},4)'    
                elif sheet_obj[f"P{i}"].value==12:
                    sheet_obj[f"K{i}"]=f"=YIELD($C$1,I{i},B{i},J{i},100,2,4)"
                    sheet_obj[f"M{i}"]=f'=EFFECT(K{i},12)'      
                else:
                    sheet_obj[f"M{i}"]=f'=YIELD(C1,I{i},B{i},J{i},100,1,0)'  
        

    #sheet_obj.delete_cols(1)
    # to give all Headings
    for i in range(3,l):
        #print("p")
        if sheet_obj[f"B{i}"].value=="GOI":
            #print("p")
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : GOI & SDL BONDS(NSDL DP Only)"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
           # sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
            sheet_obj.border=border(sheet_obj,i)
            bolt(sheet_obj,i+1)


        elif sheet_obj[f"B{i}"].value=="PSU_TAX":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : PSU Tax Free Bond"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
            
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="PSU_PERPUTUAL":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : PSU Perpetual Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="PSU_BOUND":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : PSU  Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")    
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="State_Guaranteed":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : State Guaranteed Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid") 
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="Private_Sector_AAA":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : Private Sector AAA"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="Private_Sector_Bond":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : Private Sector Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid") 
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)

        elif sheet_obj[f"B{i}"].value=="Private_Sector_Bonds":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : Private Sector Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid") 
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)    

        elif sheet_obj[f"B{i}"].value=="Private_Sector_Perpetual_Bonds":
            sheet_obj.merge_cells(f'B{i}:O{i}') 
            sheet_obj[f"B{i}"]="Category : Private Sector Perpetual Bonds"
            sheet_obj[f'B{i}'].font = Font(size=16,bold=True)
            sheet_obj[f"B{i}"].alignment = Alignment(horizontal='center', vertical='center')
            sheet_obj[f"B{i}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid") 
            #sheet_obj.row_dimensions[f"B{i}"].height = 20
            sheet_obj.border=border(sheet_obj,i)
            heading(sheet_obj,i+1)
            bolt(sheet_obj,i+1)


        

        else:
            pass   

    
    sheet_obj.merge_cells(f'B{l+1}:O{l+1}') 
    sheet_obj[f"B{l+1}"]="Please note the above rates are subject to market Fluctuations. Confirm availablity of stock and price before any Confirmation."
    sheet_obj[f"B{l+1}"].alignment = Alignment(horizontal='center', vertical='center') 
    sheet_obj[f"B{l+1}"].fill=PatternFill(start_color="00FFFF00", end_color="00FFFF00",fill_type = "solid")
    #sheet_obj.row_dimensions[f"B{l+2}"].height = 20
    sheet_obj[f'B{l+1}'].font = Font(size=14,bold=True,color='ff0000')
    i=l+1
    sheet_obj.border=border(sheet_obj,i)

    #sheet_obj.delete_cols(16)
    #Hide coloum from excelfile
    sheet_obj.column_dimensions["A"].hidden= True
    sheet_obj.column_dimensions["P"].hidden= True

    #Removing string from first row
    a=['B','D','E','F','G','H','I','J','K','L','M','N','O']
    for i in a:
        sheet_obj[f"{i}1"]=" "

    #for all test in center 
    center=['B','C','D','E','F','G','H','I','J','K','L','M','N','O']
    for j in center:
        for i in range(1,l):
            if sheet_obj[f"{j}{i}"]:
                sheet_obj[f"{j}{i}"].alignment = Alignment(horizontal='center', vertical='center') 


        

    #Replecing 0 value to NA
    for i in range(1,l):
        if sheet_obj[f"G{i}"].value==0 :
           sheet_obj[f"G{i}"].value="NA" 
        if sheet_obj[f"I{i}"].value==0:
            sheet_obj[f"I{i}"].value="NA" 
        if sheet_obj[f"K{i}"].value==0:
            sheet_obj[f"K{i}"].value="NA" 
        if sheet_obj[f"C{i}"].value:
            if sheet_obj[f"L{i}"].value is None or sheet_obj[f"L{i}"].value==0:
               sheet_obj[f"L{i}"].value="NA" 
            if sheet_obj[f"M{i}"].value is None or sheet_obj[f"M{i}"].value==0:
               sheet_obj[f"M{i}"].value="NA"   
        else:
            pass    

        
    #print(sheet_obj[f"L{i}"].value)   
    
    #change date format
    cell = sheet_obj['C1']
    cell.number_format = 'DD/mmmm/YYYY'

    for i in range(1,l):
        if type(sheet_obj[f"G{i}"].value)==datetime.datetime:
            cell = sheet_obj[f'G{i}']
            cell.number_format = 'DD/mmmm/YYYY'
        if type(sheet_obj[f"I{i}"].value)==datetime.datetime:
            cell = sheet_obj[f'I{i}']
            cell.number_format = 'DD/mmmm/YYYY'
       
    
    return wb_obj.save(filename="hello_world.xlsx")

'''def Upload(request):
        return render(request, "Upload")'''



#boredr for Heading

def border(sheet_obj,b):
        a=['B','C','D','E','F','G','H','I','J','K','L','M','N','O']
        for i in a:
            sheet_obj[f"{i}{b}"].border=Border(left=Side(border_style='thick', color='FF000000'),
            top=Side(border_style="thick", color='FF000000'),
            bottom=Side(border_style="thick", color='FF000000'),
            right=Side(border_style='thick', color='FF000000'))
        sheet_obj[f"O{b}"].border=Border(top=Side(border_style="thick", color='FF000000'),
            bottom=Side(border_style="thick", color='FF000000'),right=Side(border_style='thick', color='FF000000'))

#boredr for coloum name Heading

def border1(sheet_obj,border):
        a=['B','C','D','E','F','G','H','I','J','K','L','M','N','O']
        for i in a:
            sheet_obj[f"{i}{border}"].border=Border(left=Side(border_style='thin', color='FF000000'),
            top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),
            right=Side(border_style='thin', color='FF000000'))
        sheet_obj[f"O{border}"].border=Border(top=Side(border_style="thin", color='FF000000'),
            bottom=Side(border_style="thin", color='FF000000'),right=Side(border_style='thin', color='FF000000'))

#creading heading for all categary

def heading(sheet_obj,heading):
        a=['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
        for i in a:
            sheet_obj[f'{i}{heading}']=sheet_obj[f"{i}4"].value

#for creating to do all hrading to bolt form 

def bolt(sheet_obj,bolt):
    a=['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
    for i in a:
        sheet_obj[f'{i}{bolt}'].font = Font(bold=True)
        border1(sheet_obj,bolt)


#find user ip address

def get_ip_address(request):
    user_ip_address = request.META.get('HTTP_X_FORWARDED_FOR')
    if user_ip_address:
        ip = user_ip_address.split(',')[0]
        #print(ip)
    else:
        ip = request.META.get('REMOTE_ADDR')
        #print(ip)
    return ip

from members.models import Person1

def show_ip_address(request):
    now = datetime.datetime.now()
    dt_string = now.strftime("%Y-%m-%d %H:%M:%S")

    user_ip = get_ip_address(request)
#    print(type(user_ip))
    filename = 'ip_record.txt'
    with open(filename, 'a') as file:
        file.write(f"{user_ip} : {dt_string}" + "\n")
        person = Person1(ip=user_ip, date=dt_string)
        person.save()
    return render(request, "members/output.html", {"user_ip":user_ip})



'''def total_present_value(face_value, coupon, periods, rate):
            total_pv = 0
            for n in range(1, periods+1):
                total_pv += coupon / math.pow((1 + rate), n)
            total_pv += face_value / math.pow((1 + rate), periods)
            return total_pv'''
'''def simple(request):
    if request.method=="POST":
        person_resource=PersonReource()
        dataset=Dataset()
        now=datetime.datetime.now()
        current_date=now.strftime('%Y-%m-%d')
        
        new_person=request.FILES['C:\project\website\hello_world.xlsx']
        # dataframe=pd.read_excel(new_person)

        # dataframe.to_excel("c:\\project\\website\\webapp\\static\\upload\\{}_{}".format(current_date,new_person))
        #print('11')
        if not new_person.name.endswith("xlsx"):
            messages.info(request,"wrong format")
            return render(request,'uploade.html')
        #print('1')
        imported_data=dataset.load(new_person.read(),format='xlsx')
        #print('4')
        #print(imported_data)
        Productfield.objects.all().delete()
        for data in imported_data:
            #print(data)
            value=Productfield(
                data[0],
                data[1],
                data[2],
                data[3],
                data[4],
                data[5],
                data[6],
                data[7],
                data[8],
                data[9],
                data[10],
                data[11],
                data[12],
                data[13]
                #data[14]
                )
            value.save()
        
        dataframe=pd.read_excel(new_person)
        dataframe.to_excel("c:\\project\\website\\webapp\\static\\upload\\{}_{}".format(current_date,new_person))

        Products = Productfield.objects.all()

        i=0
        for product in Products:
            p=Products[i].Price_Per_100
            f=Products[i].Face_Value
            f=int(f)
            c=Products[i].Coupon_Rate*f
            #print(c)
            y=Products[i].Name_of_the_security
            y=int(y)

            current_date=now.strftime('%Y')
            current_date=int(current_date)
            #print(current_date)
            y=y-current_date

            ytm = ((c + (f-p)/ y)) / ((f + p)/2)
            Products[i].YTM=float(ytm*100)
            Products[i].Price_Per_100=float(product.Price_Per_100)+float(0.50)


            #print("The YTM for this bond is: " + str(ytm*100) + "%")
            i=i+1
            p=Products[i].Price_Per_100
            f=Products[i].Face_Value
            c=Products[i].Coupon_Rate
            y=Products[i].Name_of_the_security
            print(p)
            print(f)
            print(c)
            print(y)
            parser = argparse.ArgumentParser()

            parser.add_argument('-p','--Price_Per_100',  type=float,default=p, help='specifies the current price')
            #print("p")
            parser.add_argument('-f','--Face_Value',  type=float, default=f, help='specifies the face Value')
            parser.add_argument('-r','--Coupon_Rate',  type=float, default=c, help='specifies the annual coupon rate in %')
            parser.add_argument('-y','--Name_of_the_security',  type=int, default=y, help='specifies the number of years remaining to maturity')
            parser.add_argument('-s','--website',action='store_true', default=False, help='coupon is a semi-annual coupon. Default is annual')
            #print("p")
            args = parser.parse_args()
            print("p")
            coupon_rate = args.r 
            coupon = args.f * coupon_rate
            factor = 2 if args.s else 1
 
            ytm = coupon_rate
            condition = True
            while condition:
                if (args.p < args.f):
                   ytm += 0.00001
                else:
                   ytm -= 0.00001
 
                total_pv = total_present_value(args.f, coupon/factor, args.y*factor, ytm/factor)
 
                if (args.p < args.f):
                   condition = total_pv > args.p
                else:
                   condition = total_pv < args.p
            Products[i].YTM=float(ytm*100)
            i=i+1
            #print("p")'''

    
    